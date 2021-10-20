#!/usr/bin/python
# -*- coding: UTF-8 -*-
"""
@author:ZXT
@file:All_Common.py
@time:2021/06/30
"""

import requests, json, time,pymysql,sys,yaml,datetime,eventlet
from openpyxl import Workbook,load_workbook
from Config import test_host_cfg


class Common_Parameter():
    def __init__(self):
        self.host=test_host_cfg.test_host
        self.HT_host=test_host_cfg.test_HT_host

    # 根据self.host取的环境，来切换现在的数据库的环境
    def get_mysql_host(self):
        mysql_host = "h2.haier-ioc.com:3306"
        mysql_uesr = 'developer'
        mysql_password = "developer2020.YStest"
        mysql_db = "cloud_stall_test"
        mysql_port = 3306
        # 数据库连接，返回连接对象
        conn = pymysql.connect(host=mysql_host, user=mysql_uesr, password=mysql_password, db=mysql_db, port=mysql_port)

        return conn




    # 所有sql语句的公用模块
    def all_sql(self, sql_str):
        conn = self.get_mysql_host()
        cursor = conn.cursor()
        cursor.execute(sql_str)
        conn.commit()
        r = cursor.fetchall()
        cursor.close()
        conn.close()
        return r


    #获取项目的绝对路劲
    def projects_path(self):
        projects_path=sys.path[1]
        return projects_path




    #获取yml文件上面所有的文件的相对路径
    def get_yml_vlues(self):
        projects_path=self.projects_path()
        path_dict=yaml.safe_load(open(projects_path+r"\Config\file_path.yml","r",encoding="utf-8"))
        return path_dict


    #获取测试报告的绝对路劲
    def get_report_path(self):
        report_path=self.projects_path()+self.get_yml_vlues()["report_path"]
        return report_path

    #获取脚本日志文件的绝对路径
    def get_log_path(self):
        log_path=self.projects_path()+self.get_yml_vlues()["log_path"]
        return log_path

    #获取test_case文件的绝对路径
    def get_test_case_path(self):
        test_path = self.projects_path() + self.get_yml_vlues()["test_cast_path"]
        return test_path


    # 该模块的请求公用方法,返回请求地址，响应码，接口数据，接口响应的时间
    def all_requests(self,method, api, headers, **kwargs):
        """
            : method: 请求方法
            : url: 请求地址
            : kwargs: 提交参数
            :return: 返回响应数据
            """



        url =self.host + api
        try:
            with eventlet.Timeout(10):
                # 耗时操作，如果等待时间超过10秒，则直接抛出异常
                if method.upper() == "GET":
                    params=kwargs["params"]
                    r = requests.request(method, url, headers=headers, params=params)
                elif method.upper() == "POST":
                    data = json.dumps(kwargs["data"])
                    r = requests.request(method, url, headers=headers, data=data)

                data = kwargs
                code_res = r.status_code
                json_res = r.json()
                time_res = r.elapsed.total_seconds()
                return url, code_res, data, json_res, time_res

        except eventlet.timeout.Timeout:
            # 对异常进行捕获
            print("接口请求超时,超过10秒,接口请求地址：%s" % url)
            return url


    # 该模块的请求公用方法,返回请求地址，响应码，接口数据，接口响应的时间
    def all_HT_requests(self,method, api, headers, **kwargs):
        """
            : method: 请求方法
            : url: 请求地址
            : kwargs: 提交参数
            :return: 返回响应数据
            """

        url =self.HT_host + api
        try:
            with eventlet.Timeout(10):
                # 模拟耗时操作，如果等待时间超过10秒，则直接抛出异常
                r = requests.request(method, url, headers=headers, **kwargs)
                data = kwargs
                code_res = r.status_code
                json_res = r.json()
                time_res = r.elapsed.total_seconds()
                return url, code_res, data, json_res, time_res

        except eventlet.timeout.Timeout:
            # 对异常进行捕获
            print("接口请求超时,超过10秒,接口请求地址：%s" % url)
            return url



    #获取excel表格的绝对路径
    def get_excel_path(self,type="YDK"):
        f_path="excel_path_"+type
        excel_path=self.projects_path() + self.get_yml_vlues()[f_path]
        return excel_path



    #公用读取excel文档的方法，把指定的文档的数据全部读取出来，然后返回一个列表,excel_path是excel_path的文件路径，wk_name是要获取的文件的名称
    def get_excel(self,excel_path,wk_name):

        wk=load_workbook(excel_path)
        # print(wk.sheetnames)  #获取该excel文件中的所有的工作表

        #获取工作表
        sheet1=wk[wk_name]

        #获取单元格 row为行数 ，column是列数
        # print(sheet1.cell(row=2,column=1))

        #获取行数和列数
        max_row=sheet1.max_row
        max_column=sheet1.max_column

        list=[]

        for x in range(2,max_row+1):#从表格的第二行开始读取
            dict={}
            for y in range(1,max_column+1):
               key=sheet1.cell(row=1,column=y).value
               value=sheet1.cell(row=x,column=y).value
               dict[key]=value

            list.append(dict)

        return list



    #读取文档里面的值
    def get_excel_key(self,excel_path,wk_name):
        wk = load_workbook(excel_path)
        # print(wk.sheetnames)  #获取该excel文件中的所有的工作表

        # 获取工作表
        sheet1 = wk[wk_name]

        # 获取单元格 row为行数 ，column是列数
        # print(sheet1.cell(row=2,column=1))

        # 获取行数和列数
        # max_row = sheet1.max_row
        max_column = sheet1.max_column

        list = []

        for y in range(1, max_column + 1):
            key = sheet1.cell(row=1, column=y).value
            list.append(key)
        return list


    #获取现在的时间戳，然后转换成年月日，时分秒
    def get_now_time(self):
        now_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return now_time



if __name__ == '__main__':
    run=Common_Parameter()
    print(run.projects_path())

    # run.projects_path()
    # print(run.get_excel_key(r"C:\Users\Administrator\PycharmProjects\TianShi_Projects\Excel_Page\test_case_YDK.xlsx","login"))