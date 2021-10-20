#!/usr/bin/python
# -*- coding: UTF-8 -*-
"""
@author:ZXT
@file:test_case_GYS.py
@time:2021/07/01
"""
'''这个是供应商的接口自动化测试用例'''


import requests, json, time
from ddt import ddt,data,file_data,unique
import unittest
from Common_Module import All_Common
from API import login_api
from openpyxl import Workbook,load_workbook

excel_path=All_Common.Common_Parameter().get_excel_path(type="GYS")
run_get_vules=All_Common.Common_Parameter()
Supplier_vules=run_get_vules.get_excel(excel_path,"Supplier")


w_n_l=load_workbook(excel_path)

num_list=[]
for num in w_n_l.sheetnames:
    all_vules=run_get_vules.get_excel(excel_path,num)
    num_list.append(all_vules)

# print(num_list)

# for new in num_list:
#     print(new)


@ddt
class Test_Case_GYS(unittest.TestCase):
    def setUp(self):
        self.request=All_Common.Common_Parameter()
        self.token=login_api.Login_class().login(type="GYS")
        self.headers={"Authorization":self.token}




    #公用测试用例模块
    def Case_module(self,r,QW_status_code='',QW_Code=''):
        if r[1] >= 500:
            # 监控服务器是否正常
            msg = "服务器错误，响应码为：%s," \
                  "接口url:%s,请求参数为：%s" \
                  % (r[1], r[0], r[2])
            self.assertEqual(QW_status_code, r[1], msg=msg)

        elif r[4] >= 10:
            # 监控接口响应时间慢于10秒
            print("响应的时间大于10秒")
            msg = "接口响应的时间大于10秒，接口响应时间为：%s，" \
                  "接口url:%s,请求参数为：%s" % (r[4], r[0], r[2])
            self.assertEqual(3, r[4], msg=msg)

        elif r[3]["code"] !=200:
            # 正常联通测试，监控里面返回的数据是否正确，返回的数据里面code=0
            msg = "接口返回数据异常，接口url:%s,请求参数为：%s，" \
                  "接口返回的数据为%s" % (r[0], r[2], r[3])
            self.assertEqual(QW_Code, r[3]["code"], msg=msg)

        else:
            # 服务器、响应时间、接口数据都正常
            msg = "服务器、接口响应时间、接口返回数据都正常"
            self.assertEqual(QW_status_code, r[3]["code"], msg=msg)



    @data(*num_list)
    def test_022(self, num_list):
        for num in num_list:
            print(num)
            dict={}
            for num_02 in num:
                if num_02=="api":
                    api=num[num_02]
                    print("api是：",api)
                elif num_02=="QW_status_code":
                    QW_status_code=num[num_02]
                    print("响应码是：",QW_status_code)
                elif num_02=="QW_Code":
                    QW_Code=num[num_02]
                    print("预期接口响应code：",QW_Code)
                else:
                    dict[num_02]=num[num_02]

            # time.sleep(10000)
            headers = {"source": "3",
                       "Authorization": self.token}
            r = self.request.all_requests(method="GET", api=api, headers=headers, params=dict)
            print(r)
            self.Case_module(r=r, QW_status_code=QW_status_code, QW_Code=QW_Code)


    # @data(*Supplier_vules)
    # def test_01(self,kwargs):
    #     pageSize=kwargs["pageSize"]
    #     pageNum=kwargs["pageNum"]
    #     orderStatus=kwargs["orderStatus"]
    #     QW_status_code = kwargs["QW_status_code"]
    #     QW_Code = kwargs["QW_Code"]
    #     api=kwargs["api"]
    #     headers={"source":"3",
    #              "Authorization":self.token}
    #     params="pageSize={}&pageNum={}&orderStatus={}".format(pageSize,pageNum,orderStatus)
    #     r=self.request.all_requests(method="GET",api=api,headers=headers,params=params)
    #     print(r)
    #     self.Case_module(r=r,QW_status_code=QW_status_code,QW_Code=QW_Code)
#
#
if __name__ == '__main__':
    unittest.main()
