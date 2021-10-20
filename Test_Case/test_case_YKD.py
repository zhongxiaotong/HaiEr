#!/usr/bin/python
# -*- coding: UTF-8 -*-
"""
@author:ZXT
@file:test_case_YKD.py
@time:2021/06/30
"""
'''这个是云档口的接口自动化测试用例'''


import requests, json, time
from ddt import ddt,data,file_data,unique
import unittest
from Common_Module import All_Common
from API import login_api
from openpyxl import Workbook,load_workbook


excel_path=All_Common.Common_Parameter().get_excel_path(type="YDK")#获取云档口小程序测试用例Excel文件的目录

run_get_vules=All_Common.Common_Parameter()#导入自定义的公用方法

w_n_l=load_workbook(excel_path)#获取Excel文件中的所有的工作表

num_list=[]#弄一个空列表准备来装每个Excel文件中的每个子工作表里面的数据

#循环的把每个工作表里面的数据取出来
for num in w_n_l.sheetnames:#循环读取每个子表
    all_vules=run_get_vules.get_excel(excel_path,num)#获取当前子表格的所有数据，获取的数据是一个列表里面加字典，用一个变量名装起来
    num_list.append(all_vules)#把该子表格的所有数据用一个列表装起来


@ddt
class Test_Case_YDK(unittest.TestCase):
    def setUp(self):
        self.request=All_Common.Common_Parameter()
        #云档口登录，把云档口的token给提取出来
        self.token=login_api.Login_class().login(type="YDK")

        self.get_headers={"Authorization":self.token,
                          "source":"1"}
        self.post_headers={"Authorization":self.token,
                          "source":"1",
                           "content-type":"application/json"}



    #公用测试用例模块
    def Case_module(self,r,QW_status_code='',QW_Code=''):
        if r[1] >= 500:
            # 监控服务器是否正常,响应码检验
            msg = "服务器错误，响应码为：%s," \
                  "接口url:%s,请求参数为：%s" \
                  % (r[1], r[0], r[2])
            self.assertEqual(QW_status_code, r[1], msg=msg)

        elif r[4] >= 10:
            # 监控接口响应时间慢于10秒
            print("响应的时间大于10秒")
            msg = "接口响应的时间大于10秒，接口响应时间为：%s，" \
                  "接口url:%s,请求参数为：%s" % (r[4], r[0], r[2])
            self.assertEqual(3, r[4], msg=msg)

        elif r[3]["code"] !=200:
            # 正常联通测试，监控里面返回的数据是否正确，返回的数据里面code=0
            msg = "接口返回数据异常，" \
                  "接口url:%s,请求参数为：%s，" \
                  "接口返回的数据为%s" % (r[0], r[2], r[3])
            self.assertEqual(QW_Code, r[3]["code"], msg=msg)

        else:
            # 服务器、响应时间、接口数据都正常
            msg = "服务器、接口响应时间、接口返回数据都正常"
            self.assertEqual(QW_status_code, r[3]["code"], msg=msg)



    @data(*num_list)#继承Excel表格里面的所有的数据，数据是一个列表
    def test_ydk(self, num_list):
        for num in num_list:#每个子表格按照循环读出来,注意：num是字典
            print(num)
            dict={}
            for num_02 in num:#循环的读取，会把num里面所有的键 都循环读取出来
                if num_02=="api":
                    api=num[num_02]#把Excel子文档里面的api读取出来
                    print("api是：",api)
                elif num_02=="QW_status_code":
                    QW_status_code=num[num_02]#把Excel子文档里面的期望响应码读取出来
                    print("响应码是：",QW_status_code)
                elif num_02=="QW_Code":#把Excel子文档里面的期望相应code读取出来
                    QW_Code=num[num_02]
                    print("预期接口响应code：",QW_Code)

                elif num_02=="method":#把Excel子文档里面的请求方式读取出来
                    method=num[num_02]
                    print("请求方式为：",method)
                else:
                    dict[num_02]=num[num_02]#出了以上几种特殊的参数以外，其他的参数都作为请求参数传递，用字典装起来

            if method.upper() == "GET":
                headers=self.get_headers
                r = self.request.all_requests(method=method.upper(), api=api, headers=headers, params=dict)
            elif method.upper() == "POST":
                headers=self.post_headers
                r = self.request.all_requests(method=method.upper(), api=api, headers=headers, data=dict)


            self.Case_module(r=r, QW_status_code=QW_status_code, QW_Code=QW_Code)





if __name__ == '__main__':
    unittest.main()